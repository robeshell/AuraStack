# -*- coding: utf-8 -*-
"""
表格文件读写工具（CSV / XLS / XLSX）
"""
from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Iterable, List, Tuple

from flask import Response


SUPPORTED_TABLE_FILE_TYPES = ('csv', 'xls', 'xlsx')

_MIME_MAP = {
    'csv': 'text/csv; charset=utf-8',
    'xls': 'application/vnd.ms-excel',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}


def normalize_table_file_type(raw_value, default='csv'):
    value = str(raw_value or '').strip().lower()
    if value in SUPPORTED_TABLE_FILE_TYPES:
        return value
    return default


def infer_table_file_type(filename):
    name = str(filename or '').strip().lower()
    if not name or '.' not in name:
        return None
    ext = name.rsplit('.', 1)[-1]
    if ext in SUPPORTED_TABLE_FILE_TYPES:
        return ext
    return None


def _format_cell_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _format_xls_cell(cell_type, cell_value, datemode):
    if cell_value is None:
        return ''
    if cell_type == 3:  # XL_CELL_DATE
        try:
            import xlrd
            value = xlrd.xldate_as_datetime(cell_value, datemode)
            return value.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return ''
    if isinstance(cell_value, float) and cell_value.is_integer():
        return str(int(cell_value))
    return str(cell_value).strip()


def _clean_header_row(headers: Iterable):
    cleaned = []
    for header in headers:
        cleaned.append(str(header or '').strip())
    return cleaned


def _collect_row_map(headers, values):
    row_map = {}
    has_data = False
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else ''
        text = _format_cell_value(value)
        row_map[header] = text
        if text != '':
            has_data = True
    return row_map, has_data


def read_table_file(file_storage):
    """
    读取上传文件并返回:
    - fieldnames: 表头数组
    - rows_with_line: [(line_number, row_dict), ...]
    - file_type: csv/xls/xlsx
    """
    if not file_storage:
        raise ValueError('请上传导入文件')

    file_type = infer_table_file_type(file_storage.filename)
    if not file_type:
        raise ValueError('仅支持 csv/xls/xlsx 文件')

    content = file_storage.read()
    if not content:
        raise ValueError('导入文件内容为空')

    if file_type == 'csv':
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise ValueError('CSV 编码错误，请使用 UTF-8 编码') from exc

        reader = csv.DictReader(StringIO(text))
        fieldnames = [str(item or '') for item in (reader.fieldnames or [])]
        rows = []
        line = 1
        for row in reader:
            line += 1
            normalized_row = {k: ('' if v is None else str(v).strip()) for k, v in (row or {}).items()}
            if not any(str(value).strip() for value in normalized_row.values()):
                continue
            rows.append((line, normalized_row))
        return fieldnames, rows, file_type

    if file_type == 'xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError('缺少 openpyxl 依赖，无法处理 xlsx 文件') from exc

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], [], file_type

        headers = _clean_header_row([_format_cell_value(item) for item in header_row])
        rows = []
        for line_number, row_values in enumerate(rows_iter, start=2):
            values = list(row_values or [])
            row_map, has_data = _collect_row_map(headers, values)
            if has_data:
                rows.append((line_number, row_map))
        return headers, rows, file_type

    # xls
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError('缺少 xlrd 依赖，无法处理 xls 文件') from exc

    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows <= 0:
        return [], [], file_type

    headers = _clean_header_row([
        _format_xls_cell(sheet.cell_type(0, col), sheet.cell_value(0, col), workbook.datemode)
        for col in range(sheet.ncols)
    ])

    rows = []
    for row_index in range(1, sheet.nrows):
        raw_values = [
            _format_xls_cell(sheet.cell_type(row_index, col), sheet.cell_value(row_index, col), workbook.datemode)
            for col in range(sheet.ncols)
        ]
        row_map, has_data = _collect_row_map(headers, raw_values)
        if has_data:
            rows.append((row_index + 1, row_map))
    return headers, rows, file_type


def build_table_response(headers: List[str], rows: Iterable[Iterable], base_filename: str, file_type='csv'):
    file_type = normalize_table_file_type(file_type, default='csv')
    safe_headers = [str(item) for item in (headers or [])]
    safe_rows = [[_format_cell_value(value) for value in (row or [])] for row in rows]

    if file_type == 'csv':
        stream = StringIO()
        writer = csv.writer(stream)
        writer.writerow(safe_headers)
        for row in safe_rows:
            writer.writerow(row)
        payload = ('\ufeff' + stream.getvalue()).encode('utf-8')
    elif file_type == 'xlsx':
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError('缺少 openpyxl 依赖，无法导出 xlsx 文件') from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(safe_headers)
        for row in safe_rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()
    else:
        try:
            import xlwt
        except ImportError as exc:
            raise RuntimeError('缺少 xlwt 依赖，无法导出 xls 文件') from exc

        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Sheet1')
        for col, value in enumerate(safe_headers):
            sheet.write(0, col, value)
        for row_index, row in enumerate(safe_rows, start=1):
            for col_index, value in enumerate(row):
                sheet.write(row_index, col_index, value)
        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()

    response = Response(payload, mimetype=_MIME_MAP[file_type])
    response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.{file_type}'
    return response
